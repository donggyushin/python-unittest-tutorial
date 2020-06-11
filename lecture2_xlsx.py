import openpyxl as xr
import xlsxwriter as xw


def read_xlsx(filePath, sheetName=None):
    book = xr.load_workbook(filePath, data_only=True)
    return book, (book[sheetName] if sheetName else book.active)


def _write_xlsx(filePath, sheetName, data):
    book = xw.Workbook(filePath)
    sheet = book.add_worksheet(name=sheetName)

    row = 0
    col = 0

    for r in data:
        for c in r:
            sheet.write(row, col, c)
            col += 1

        row += 1
        col = 0

    return book, sheet


def write_xlsx(filePath, sheetName, data):
    book, sheet = _write_xlsx(filePath, sheetName, data)
    book.close()


def white_format_xlsx(filePath, sheetName, data, cell, format, condition, bold=True):
    book, sheet = _write_xlsx(filePath, sheetName, data)

    fm = book.add_format(format)
    if bold:
        fm.set_bold()
    # condition['format'] == fm
    sheet.conditional_format(cell, condition)

    book.close()


def write_chart_xlsx(filePath, sheetName, data, cells, chart, series):
    book = xw.Workbook(filePath)
    sheet = book.add_worksheet(name=sheetName)

    sheet.write_column('A1', data)

    chart = book.add_chart(chart)
    chart.add_series(series)

    sheet.insert_chart(cells, chart)

    book.close()


def financial_analysis(filePath):
    book = xw.Workbook(filePath)
    sheet = book.add_worksheet()

    chart = book.add_chart({'type': 'column'})
    data = [
        [
            'Year', '2013', '2014', '2015'
        ],
        [
            'Revenue', 100, 120, 125
        ],
        [
            'COGS', 80, 90, 70
        ],
        [
            'Profit', 20, 30, 55
        ]
    ]

    sheet.write_row('A1', data[0])
    sheet.write_row('A2', data[1])
    sheet.write_row('A3', data[2])
    sheet.write_row('A4', data[3])

    chart.add_series({'values': '=Sheet1!$B$2:$B$4'})
    chart.add_series({'values': '=Sheet1!$C$2:$C$4'})
    chart.add_series({'values': '=Sheet1!$D$2:$D$4'})
    sheet.insert_chart('G1', chart)

    sheet.write(5, 0, '% Gain')
    sheet.write(5, 1, '=(B4/B2)*100')
    sheet.write(5, 2, '=(C4/C2)*100')
    sheet.write(5, 3, '=(D4/D2)*100')

    book.close()
